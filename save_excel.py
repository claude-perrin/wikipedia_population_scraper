import openpyxl


def open_excel_file(path='/python/Upwork/Wikipedia/wir-machen-druck_Briefpapier_Demo.xlsx'):
    wb_obj = openpyxl.load_workbook(path)
    return wb_obj, wb_obj.active


def find_last_empty_cell():
    column = open_excel_file()[1]['A']
    last_empty_cell_index = len(column)
    for cell in column:
        if cell.value is None:
            last_empty_cell_index = cell.row
            break

    if last_empty_cell_index == 1:
        return last_empty_cell_index - 1
    else:
        return last_empty_cell_index


def insert_word_into_excel(category, keywords):
    empty_cell = find_last_empty_cell()
    wb_sheet_obj = open_excel_file()
    y = 0
    for key in keywords:
        print(keywords[key])
        wb_sheet_obj[1].cell(empty_cell + 1, 1).value = f'Wikipedia {category["type"]}'
        wb_sheet_obj[1].cell(empty_cell + 1, 2).value = f'https://en.wikipedia.org/wiki/{category["country"]}'
        try:
            wb_sheet_obj[1].cell(empty_cell + 1, 3).value = key
            wb_sheet_obj[1].cell(empty_cell + 1, 4).value = keywords[key]
        except ValueError:
            wb_sheet_obj[1].cell(empty_cell + 1, 3).value = key
            value = ''
            for i in keywords[key]:
                value += i + ' '
            wb_sheet_obj[1].cell(empty_cell + 1, 4).value = value

        empty_cell += 1
        y += 1
    wb_sheet_obj[0].save("wir-machen-druck_Briefpapier_Demo.xlsx")


if __name__ == '__main__':
    pass
